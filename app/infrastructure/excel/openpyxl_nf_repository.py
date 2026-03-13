from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from config.settings import EXCEL_FILE, EXPECTED_COLUMNS, TABLE_NAME, WORKSHEET_NAME
from app.domain.entities.nf_item import NfItem
from app.domain.interfaces.excel_repository import ExcelRepository


class OpenpyxlNfRepository(ExcelRepository):
    def get_existing_nfs(self) -> set[str]:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[WORKSHEET_NAME]

        if TABLE_NAME not in ws.tables:
            raise ValueError(f"Tabela '{TABLE_NAME}' não encontrada na aba '{WORKSHEET_NAME}'.")

        tabela = ws.tables[TABLE_NAME]
        start_cell, end_cell = tabela.ref.split(":")
        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)

        start_col_idx = column_index_from_string(start_col_letter)
        end_col_idx = column_index_from_string(end_col_letter)

        header_map = {}
        for col in range(start_col_idx, end_col_idx + 1):
            value = ws.cell(row=start_row, column=col).value
            if value is not None:
                header_map[str(value).strip()] = col

        if "Numero NF" not in header_map:
            raise ValueError("Não encontrei a coluna 'Numero NF' na tabela.")

        col_nf = header_map["Numero NF"]

        existentes = set()
        for row in range(start_row + 1, end_row + 1):
            value = ws.cell(row=row, column=col_nf).value
            if value is not None:
                nf = str(value).strip()
                if nf:
                    existentes.add(nf)

        return existentes

    def insert_items(self, items: list[NfItem]) -> int:
        if not items:
            return 0

        wb = load_workbook(EXCEL_FILE)
        ws = wb[WORKSHEET_NAME]

        if TABLE_NAME not in ws.tables:
            raise ValueError(f"Tabela '{TABLE_NAME}' não encontrada na aba '{WORKSHEET_NAME}'.")

        tabela = ws.tables[TABLE_NAME]
        start_cell, end_cell = tabela.ref.split(":")
        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)

        start_col_idx = column_index_from_string(start_col_letter)
        end_col_idx = column_index_from_string(end_col_letter)

        header_map = {}
        for col in range(start_col_idx, end_col_idx + 1):
            value = ws.cell(row=start_row, column=col).value
            if value is not None:
                header_map[str(value).strip()] = col

        missing = [column for column in EXPECTED_COLUMNS if column not in header_map]
        if missing:
            raise ValueError("Faltam cabeçalhos na tabela: " + ", ".join(missing))

        insert_at = end_row + 1
        ws.insert_rows(insert_at, amount=len(items))

        current_row = insert_at
        for item in items:
            row_values = dict(zip(EXPECTED_COLUMNS, item.to_row()))
            for column_name in EXPECTED_COLUMNS:
                col_index = header_map[column_name]
                ws.cell(row=current_row, column=col_index, value=row_values[column_name])
            current_row += 1

        new_end_row = end_row + len(items)
        tabela.ref = f"{start_col_letter}{start_row}:{end_col_letter}{new_end_row}"

        wb.save(EXCEL_FILE)
        return len(items)